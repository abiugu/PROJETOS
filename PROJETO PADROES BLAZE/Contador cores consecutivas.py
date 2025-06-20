import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Função para determinar a cor de um número, dado o número da célula anterior
def cor_do_numero(num):
    """Determina a cor com base no número da célula anterior."""
    if num == 0:
        return "white"  # Branco
    elif 1 <= num <= 7:
        return "red"    # Vermelho
    elif 8 <= num <= 14:
        return "black"  # Preto
    return None

# Função para contar as aparições de cada número, as cores da célula de cima (jogada anterior),
# o máximo de sequências consecutivas de cada cor, a quantidade de cada sequência e calcular a média das sequências
def contar_cores_jogada_anterior(df, wb):
    """Conta as aparições de cada número (0 a 14), as cores da célula de cima,
    o máximo de sequências consecutivas de cada cor, a quantidade de cada sequência, a média das sequências e os percentuais."""
    
    resultado = {}
    sheet = wb.active  # Seleciona a planilha ativa (você pode mudar isso para uma planilha específica)

    # Iterar sobre todos os números de 0 a 14
    for numero in range(0, 15):  # Agora percorre todos os números de 0 a 14
        jogadas_numero = df[df['Número'] == numero]
        
        # Inicializa a contagem das cores da célula de cima para cada número
        contagem_cor_anterior = {'white': 0, 'red': 0, 'black': 0}
        
        # Contador de aparições do número
        contagem_total = len(jogadas_numero)
        
        # Inicializa as variáveis para contar a sequência máxima de cores consecutivas
        sequencia_max = {'white': 0, 'red': 0, 'black': 0}  # Sequência máxima
        sequencia_atual = {'white': 0, 'red': 0, 'black': 0}  # Sequência atual
        linha_sequencia_max = {'white': None, 'red': None, 'black': None}  # Linha da maior sequência
        
        # Contadores para as quantidades de cada sequência de cores
        contagem_sequencias = {'white': {}, 'red': {}, 'black': {}}  # Dicionário para contar os comprimentos das sequências

        # Dicionário para contar as sequências seguintes após cada tipo de sequência
        sequencias_seguintes = {'white': {}, 'red': {}, 'black': {}}  # Ex: Após sequência de 3 cores, quantas vezes 1, 2, 3 cores apareceram
        
        cor_anterior = None  # Variável para armazenar a cor anterior
        linha_anterior = None  # Variável para armazenar a linha da célula de cima
        
        # Primeira contagem: Contagem das sequências principais
        for i in range(1, len(jogadas_numero)):  # Inicia de 1 para acessar a célula de cima
            # Acessa a célula de cima (linha acima)
            celula_acima = sheet.cell(row=jogadas_numero.iloc[i].name, column=2)  # Coluna 2 é a "Cor"
            
            # Determina a cor da célula de cima
            cor_atual = cor_do_numero(df.iloc[jogadas_numero.iloc[i].name - 1]['Número'])  # A célula de cima
            
            # Verifica se a cor é válida
            if cor_atual:
                contagem_cor_anterior[cor_atual] += 1
                
                # Se a cor for a mesma que a anterior, aumenta a sequência
                if cor_atual == cor_anterior:
                    sequencia_atual[cor_atual] += 1
                else:
                    # Se a cor mudou, reseta a sequência para a nova cor
                    if cor_anterior:
                        # Atualiza a sequência máxima da cor anterior, caso a sequência atual seja maior
                        if sequencia_atual[cor_anterior] > sequencia_max[cor_anterior]:
                            sequencia_max[cor_anterior] = sequencia_atual[cor_anterior]
                            linha_sequencia_max[cor_anterior] = linha_anterior
                        
                        # Conta a sequência de comprimento da cor anterior
                        seq_len = sequencia_atual[cor_anterior]
                        if seq_len in contagem_sequencias[cor_anterior]:
                            contagem_sequencias[cor_anterior][seq_len] += 1
                        else:
                            contagem_sequencias[cor_anterior][seq_len] = 1
                    
                    sequencia_atual[cor_atual] = 1  # Reinicia a sequência para a nova cor
                    linha_anterior = jogadas_numero.iloc[i].name  # Atualiza a linha para a célula atual
                
            # Atualiza a cor anterior
            cor_anterior = cor_atual
        
        # Atualiza a sequência máxima para a última cor da iteração
        if cor_anterior and sequencia_atual[cor_anterior] > sequencia_max[cor_anterior]:
            sequencia_max[cor_anterior] = sequencia_atual[cor_anterior]
            linha_sequencia_max[cor_anterior] = linha_anterior
        
        # Conta a sequência final
        if cor_anterior:
            seq_len = sequencia_atual[cor_anterior]
            if seq_len in contagem_sequencias[cor_anterior]:
                contagem_sequencias[cor_anterior][seq_len] += 1
            else:
                contagem_sequencias[cor_anterior][seq_len] = 1
        
        # Segunda contagem: Sequências seguintes após a sequência principal
        cor_anterior = None  # Reiniciar a cor anterior para a nova iteração
        sequencia_atual = {'white': 0, 'red': 0, 'black': 0}  # Reiniciar a sequência
        
        for i in range(1, len(jogadas_numero)):  # Inicia de 1 para acessar a célula de cima
            cor_atual = cor_do_numero(df.iloc[jogadas_numero.iloc[i].name - 1]['Número'])
            
            if cor_atual:
                # Se a cor for a mesma que a anterior, aumenta a sequência
                if cor_atual == cor_anterior:
                    sequencia_atual[cor_atual] += 1
                else:
                    # Se a cor mudou, conta a sequência seguinte que apareceu
                    if cor_anterior:
                        sequencia_anterior = sequencia_atual[cor_anterior]
                        
                        # Registra a sequência seguinte
                        if sequencia_anterior not in sequencias_seguintes[cor_anterior]:
                            sequencias_seguintes[cor_anterior][sequencia_anterior] = {}
                    
                        if sequencia_atual[cor_atual] not in sequencias_seguintes[cor_anterior][sequencia_anterior]:
                            sequencias_seguintes[cor_anterior][sequencia_anterior][sequencia_atual[cor_atual]] = 1
                        else:
                            sequencias_seguintes[cor_anterior][sequencia_anterior][sequencia_atual[cor_atual]] += 1
                    
                    sequencia_atual[cor_atual] = 1  # Reinicia a sequência para a nova cor
            
            cor_anterior = cor_atual
        
        # Calcular percentuais para as sequências principais
        percentuais_sequencias = {'white': {}, 'red': {}, 'black': {}}
        for cor in ['white', 'red', 'black']:
            total_ocorrencias = sum(contagem_sequencias[cor].values())
            if total_ocorrencias > 0:
                for seq_len, count in contagem_sequencias[cor].items():
                    percentuais_sequencias[cor][seq_len] = (count / total_ocorrencias) * 100
            else:
                percentuais_sequencias[cor] = {}

        # Calcular percentuais para as sequências seguintes
        percentuais_sequencias_seguintes = {'white': {}, 'red': {}, 'black': {}}
        for cor in ['white', 'red', 'black']:
            for seq_len, subseq_dict in sequencias_seguintes[cor].items():
                total_subsequentes = sum(subseq_dict.values())
                if total_subsequentes > 0:
                    for subseq_len, count in subseq_dict.items():
                        percentuais_sequencias_seguintes[cor].setdefault(seq_len, {})[subseq_len] = (count / total_subsequentes) * 100
                else:
                    percentuais_sequencias_seguintes[cor][seq_len] = {}

        # Armazenar os resultados para o número atual
        resultado[numero] = {
            'total': contagem_total,  # Total de aparições do número
            'cores_seguinte': contagem_cor_anterior,  # Contagem das cores da célula de cima
            'sequencias_maximas': sequencia_max,  # Máximas sequências consecutivas
            'linha_sequencia_max': linha_sequencia_max,  # Linha da sequência máxima de cada cor
            'contagem_sequencias': contagem_sequencias,  # Contagem de quantas vezes cada sequência ocorre
            'percentuais_sequencias': percentuais_sequencias,  # Percentuais das sequências de cada cor
            'sequencias_seguintes': sequencias_seguintes,  # Sequências seguintes após cada sequência principal
            'percentuais_sequencias_seguintes': percentuais_sequencias_seguintes  # Percentuais das sequências seguintes
        }

    return resultado

# Caminho do arquivo Excel no Desktop
file_path = os.path.expanduser('~/Desktop/historico blaze.xlsx')  # Ajuste o nome do arquivo se necessário

# Carregar o arquivo Excel
df = pd.read_excel(file_path)

# Carregar o arquivo Excel com openpyxl para acessar as cores das células
wb = load_workbook(file_path)

# Realizar a análise para todos os números
resultado_analise = contar_cores_jogada_anterior(df, wb)

# Caminho do arquivo de saída (no Desktop)
output_file_path = os.path.expanduser('~/Desktop/resultado_analise_sequencias seguintes.txt')

# Salvar os resultados em um arquivo .txt no Desktop com formatação mais organizada
with open(output_file_path, 'w') as file:
    for numero, dados in resultado_analise.items():
        file.write(f"==============================\n")
        file.write(f"          Número {numero}          \n")
        file.write(f"==============================\n")
        file.write(f"\n")
        
        file.write(f"** Total de aparições: {dados['total']} **\n\n")

        # Contagem das cores seguintes (da célula de cima)
        file.write(f"** Cores seguintes (da célula de cima): **\n")
        for cor, quantidade in dados['cores_seguinte'].items():
            percentual = (quantidade / dados['total']) * 100
            file.write(f"  - {cor.capitalize()}: {quantidade} vezes ({percentual:.2f}%)\n")
        file.write(f"\n")

        # Máximas sequências consecutivas
        file.write(f"** Máximas sequências consecutivas: **\n")
        for cor, sequencia in dados['sequencias_maximas'].items():
            file.write(f"  - {cor.capitalize()}: {sequencia} (Linha: {dados['linha_sequencia_max'][cor]})\n")
        file.write(f"\n")

        # Contagem das sequências
        file.write(f"** Contagem das sequências principais: **\n")
        for cor, seq_dict in dados['contagem_sequencias'].items():
            file.write(f"  - {cor.capitalize()}:\n")
            for seq_len, count in sorted(seq_dict.items()):  # Ordenando as sequências por tamanho
                percentual = dados['percentuais_sequencias'][cor].get(seq_len, 0)
                file.write(f"    - Sequência de {seq_len} cores: {count} vezes ({percentual:.2f}%)\n")
            file.write(f"\n")

        # Sequências seguintes (após cada sequência principal)
        file.write(f"** Sequências seguintes após cada sequência principal: **\n")
        for cor, seq_dict in dados['sequencias_seguintes'].items():
            file.write(f"  - {cor.capitalize()}:\n")
            for seq_len, subseq_dict in sorted(seq_dict.items()):  # Ordenando as sequências por tamanho
                file.write(f"    - Após sequência de {seq_len} cores:\n")
                for subseq_len, subseq_count in sorted(subseq_dict.items()):
                    percentual = dados['percentuais_sequencias_seguintes'][cor].get(seq_len, {}).get(subseq_len, 0)
                    file.write(f"      - Sequência de {subseq_len} cores: {subseq_count} vezes ({percentual:.2f}%)\n")
            file.write(f"\n")

        file.write(f"==============================\n")
        file.write(f"\n")

print(f"Resultado salvo em: {output_file_path}")

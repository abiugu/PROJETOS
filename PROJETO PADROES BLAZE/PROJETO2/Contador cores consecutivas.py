import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Função para determinar a cor de um número, dado o número da célula anterior
def cor_do_numero(num):
    """Determina a cor com base no número da célula anterior."""
    if num == 0:
        return "white"  # Branco
    elif 1 <= num <= 7:
        return "red"    # Vermelho
    elif 8 <= num <= 14:
        return "black"  # Preto
    return None

# Função para contar as aparições de cada número, as cores da célula de cima (jogada anterior),
# o máximo de sequências consecutivas de cada cor, a quantidade de cada sequência e calcular a média das sequências
def contar_cores_jogada_anterior(df, wb):
    """Conta as aparições de cada número (0 a 14), as cores da célula de cima,
    o máximo de sequências consecutivas de cada cor, a quantidade de cada sequência, a média das sequências e os percentuais."""
    
    resultado = {}
    sheet = wb.active  # Seleciona a planilha ativa (você pode mudar isso para uma planilha específica)

    # Iterar sobre todos os números de 0 a 14
    for numero in range(0, 15):  # Agora percorre todos os números de 0 a 14
        jogadas_numero = df[df['Número'] == numero]
        
        # Inicializa a contagem das cores da célula de cima para cada número
        contagem_cor_anterior = {'white': 0, 'red': 0, 'black': 0}
        
        # Contador de aparições do número
        contagem_total = len(jogadas_numero)
        
        # Inicializa as variáveis para contar a sequência máxima de cores consecutivas
        sequencia_max = {'white': 0, 'red': 0, 'black': 0}  # Sequência máxima
        sequencia_atual = {'white': 0, 'red': 0, 'black': 0}  # Sequência atual
        linha_sequencia_max = {'white': None, 'red': None, 'black': None}  # Linha da maior sequência
        
        # Contadores para as quantidades de cada sequência de cores
        contagem_sequencias = {'white': {}, 'red': {}, 'black': {}}  # Dicionário para contar os comprimentos das sequências

        # Dicionário para contar as sequências seguintes após cada tipo de sequência
        sequencias_seguintes = {'white': {}, 'red': {}, 'black': {}}  # Ex: Após sequência de 3 cores, quantas vezes 1, 2, 3 cores apareceram
        
        cor_anterior = None  # Variável para armazenar a cor anterior
        linha_anterior = None  # Variável para armazenar a linha da célula de cima
        
        # Contagem das sequências principais
        for i in range(1, len(jogadas_numero)):  # Inicia de 1 para acessar a célula de cima
            # Acessa a célula de cima (linha acima)
            celula_acima = sheet.cell(row=jogadas_numero.iloc[i].name, column=2)  # Coluna 2 é a "Cor"
            
            # Determina a cor da célula de cima
            cor_atual = cor_do_numero(df.iloc[jogadas_numero.iloc[i].name - 1]['Número'])  # A célula de cima
            
            # Verifica se a cor é válida
            if cor_atual:
                contagem_cor_anterior[cor_atual] += 1
                
                # Se a cor for a mesma que a anterior, aumenta a sequência
                if cor_atual == cor_anterior:
                    sequencia_atual[cor_atual] += 1
                else:
                    # Se a cor mudou, reseta a sequência para a nova cor
                    if cor_anterior:
                        # Atualiza a sequência máxima da cor anterior, caso a sequência atual seja maior
                        if sequencia_atual[cor_anterior] > sequencia_max[cor_anterior]:
                            sequencia_max[cor_anterior] = sequencia_atual[cor_anterior]
                            linha_sequencia_max[cor_anterior] = linha_anterior
                        
                        # Conta a sequência de comprimento da cor anterior
                        seq_len = sequencia_atual[cor_anterior]
                        if seq_len in contagem_sequencias[cor_anterior]:
                            contagem_sequencias[cor_anterior][seq_len] += 1
                        else:
                            contagem_sequencias[cor_anterior][seq_len] = 1
                    
                    sequencia_atual[cor_atual] = 1  # Reinicia a sequência para a nova cor
                    linha_anterior = jogadas_numero.iloc[i].name  # Atualiza a linha para a célula atual
                
            # Atualiza a cor anterior
            cor_anterior = cor_atual
        
        # Atualiza a sequência máxima para a última cor da iteração
        if cor_anterior and sequencia_atual[cor_anterior] > sequencia_max[cor_anterior]:
            sequencia_max[cor_anterior] = sequencia_atual[cor_anterior]
            linha_sequencia_max[cor_anterior] = linha_anterior
        
        # Conta a sequência final
        if cor_anterior:
            seq_len = sequencia_atual[cor_anterior]
            if seq_len in contagem_sequencias[cor_anterior]:
                contagem_sequencias[cor_anterior][seq_len] += 1
            else:
                contagem_sequencias[cor_anterior][seq_len] = 1
        
        # Armazenar os resultados para o número atual
        resultado[numero] = {
            'total': contagem_total,  # Total de aparições do número
            'cores_seguinte': contagem_cor_anterior,  # Contagem das cores da célula de cima
            'sequencias_maximas': sequencia_max,  # Máximas sequências consecutivas
            'linha_sequencia_max': linha_sequencia_max,  # Linha da sequência máxima de cada cor
            'contagem_sequencias': contagem_sequencias,  # Contagem de quantas vezes cada sequência ocorre
        }

    # Agora, vamos calcular os percentuais
    for numero, dados in resultado.items():
        # Cores seguintes
        total_aparicoes = dados['total']
        percentuais_cores = {}
        for cor, contagem in dados['cores_seguinte'].items():
            percentuais_cores[cor] = (contagem / total_aparicoes) * 100
        dados['percentuais_cores'] = percentuais_cores

        # Sequências principais
        percentuais_sequencias = {}
        for cor, seq_dict in dados['contagem_sequencias'].items():
            percentuais_sequencias[cor] = {}
            total_sequencias = sum(seq_dict.values())
            for seq_len, count in seq_dict.items():
                percentuais_sequencias[cor][seq_len] = (count / total_sequencias) * 100
        dados['percentuais_sequencias'] = percentuais_sequencias

    return resultado

# Caminho do arquivo Excel no Desktop
file_path = os.path.expanduser('~/Desktop/historico blaze teste.xlsx')  # Ajuste o nome do arquivo se necessário

# Carregar o arquivo Excel
df = pd.read_excel(file_path)

# Carregar o arquivo Excel com openpyxl para acessar as cores das células
wb = load_workbook(file_path)

# Realizar a análise para todos os números
resultado_analise = contar_cores_jogada_anterior(df, wb)

# Caminho do arquivo de saída (no Desktop)
output_file_path = os.path.expanduser('~/Desktop/analise max sequencias seguintes.txt')

# Salvar os resultados em um arquivo .txt no Desktop com formatação mais organizada
with open(output_file_path, 'w') as file:
    for numero, dados in resultado_analise.items():
        file.write(f"==============================\n")
        file.write(f"          Número {numero}          \n")
        file.write(f"==============================\n")
        file.write(f"\n")
        
        file.write(f"** Total de aparições: {dados['total']} **\n\n")

        # Contagem das cores seguintes (da célula de cima)
        file.write(f"** Cores seguintes (da célula de cima): **\n")
        for cor, quantidade in dados['cores_seguinte'].items():
            percentual = dados['percentuais_cores'].get(cor, 0)
            file.write(f"  - {cor.capitalize()}: {quantidade} vezes ({percentual:.2f}%)\n")
        file.write(f"\n")

        # Máximas sequências consecutivas
        file.write(f"** Máximas sequências consecutivas: **\n")
        for cor, sequencia in dados['sequencias_maximas'].items():
            file.write(f"  - {cor.capitalize()}: {sequencia} (Linha: {dados['linha_sequencia_max'][cor]})\n")
        file.write(f"\n")

        # Contagem das sequências
        file.write(f"** Contagem das sequências principais: **\n")
        for cor, seq_dict in dados['contagem_sequencias'].items():
            file.write(f"  - {cor.capitalize()}:\n")
            for seq_len, count in sorted(seq_dict.items()):  # Ordenando as sequências por tamanho
                percentual = dados['percentuais_sequencias'][cor].get(seq_len, 0)
                file.write(f"    - Sequência de {seq_len} cores: {count} vezes ({percentual:.2f}%)\n")
            file.write(f"\n")

        file.write(f"==============================\n")
        file.write(f"\n")

print(f"Resultado salvo em: {output_file_path}")

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import threading

pares = [
('48.94', '37.5'),
('52.75', '56.25'),
('56.84', '65.31'),
('51.65', '62.22'),
('56.52', '52.27'),
('47.25', '41.67'),
('59.14', '55.56'),
('56.18', '53.49'),
('41.24', '38.78'),
('38.54', '39.58'),
('37.63', '36.17'),
('43.62', '34.78'),
('46.32', '55.1'),
('43.96', '44.19'),
('50.54', '43.18'),
('45.16', '36.36'),
('44.9', '50'),
('49.45', '54.17'),
('55.43', '47.73'),
('51.09', '40'),
('40.22', '47.83'),
('40.22', '45.65'),
('53.61', '44.68'),
('52.63', '44.44'),
('58.89', '64.44'),
('47.83', '44.19'),
('47.83', '58.7'),
('56.12', '54.17'),
('53.93', '58.7'),
('57.14', '55.81'),
('53.06', '60.42'),
('54.44', '46.51'),
('48.91', '38.64'),
('51.11', '52.08'),
('38.04', '34.78'),
('42.71', '33.33'),
('41.3', '47.73'),
('53.68', '56'),
('58.16', '54'),
('42.86', '33.33'),
('53.68', '51.11'),
('58.43', '52.27'),
('58.51', '66.67'),
('50.57', '47.73'),
('53.06', '44'),
('52.22', '44.19'),
('55.56', '56'),
('55.56', '51.06'),
('56.99', '54.55'),
('60.2', '61.22'),
('47.87', '48'),
('48.94', '54.55'),
('54.84', '65.96'),
('55.06', '57.14'),
('45.45', '43.18'),
('61.54', '61.36'),
('59.14', '51.06'),
('50.53', '62.5'),
('38.54', '38.3'),
('56.04', '46.81'),
('43.18', '46.51'),
('45.26', '55.1'),
('56.84', '62'),
('61.22', '61.22'),
('42.7', '44.68'),
('61.86', '61.22'),
('46.15', '36.17'),
('57.73', '66.67'),
('52.53', '53.06'),
('51.09', '58.14'),
('48.98', '56.25'),
('53.26', '51.02'),
('49.44', '52.38')
]

def analisar_planilha():
    caminho = filedialog.askopenfilename(
        title="Selecione a planilha Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
    )
    if not caminho:
        return

    root.destroy()

    progresso_janela = tk.Tk()
    progresso_janela.title("Processando...")
    progresso_janela.geometry("400x120")
    progresso_janela.resizable(False, False)

    tk.Label(progresso_janela, text="⏳ Analisando planilha, aguarde...",
             font=("Arial", 11)).pack(pady=10)

    barra = ttk.Progressbar(progresso_janela, orient="horizontal",
                            length=300, mode="determinate", maximum=100)
    barra.pack(pady=10)
    progresso_texto = tk.Label(progresso_janela, text="0%", font=("Arial", 10))
    progresso_texto.pack()

    def atualizar_progresso(valor):
        barra["value"] = valor
        progresso_texto.config(text=f"{int(valor)}%")
        progresso_janela.update_idletasks()

    def executar():
        try:
            df = pd.read_excel(caminho)
        except Exception as e:
            progresso_janela.after(0, lambda: [
                progresso_janela.destroy(),
                messagebox.showerror("Erro", f"Erro ao abrir o arquivo:\n{e}")
            ])
            return

        if not {'Probabilidade 100', 'Probabilidade 50', 'Cor'}.issubset(df.columns):
            progresso_janela.after(0, lambda: [
                progresso_janela.destroy(),
                messagebox.showerror("Erro", "A planilha precisa conter as colunas 'Cor', 'Probabilidade 100' e 'Probabilidade 50'.")
            ])
            return

        df['Probabilidade 100'] = df['Probabilidade 100'].astype(str).str.strip()
        df['Probabilidade 50'] = df['Probabilidade 50'].astype(str).str.strip()
        df['Cor'] = df['Cor'].astype(str).str.strip().str.lower()

        total = len(df)

        encontrados = []
        for i, row in df.iterrows():
            encontrados.append((row['Probabilidade 100'], row['Probabilidade 50']) in pares)
            if i % 5 == 0:
                progresso_janela.after(0, lambda v=(i / total) * 50: atualizar_progresso(v))

        df['Encontrado'] = encontrados

        resultados = []
        for i in range(len(df)):
            prox1 = df.loc[i + 1, 'Cor'] if i + 1 < len(df) else None
            prox2 = df.loc[i + 2, 'Cor'] if i + 2 < len(df) else None

            if prox1 == 'red':
                resultados.append('Acerto G0')
            elif prox2 == 'red':
                resultados.append('Acerto G1')
            else:
                resultados.append('Erro')

            if i % 5 == 0:
                progresso_janela.after(0, lambda v=50 + (i / total) * 40: atualizar_progresso(v))

        df['Resultado'] = resultados

        encontrados_total = df['Encontrado'].sum()
        percentual_encontrado = (encontrados_total / total * 100) if total > 0 else 0
        acertos = df['Resultado'].str.contains('Acerto').sum()
        erros = df['Resultado'].str.contains('Erro').sum()
        percentual_acerto = (acertos / total * 100) if total > 0 else 0

        salvar = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Salvar resultado como..."
        )
        if not salvar:
            progresso_janela.after(0, progresso_janela.destroy)
            return

        df.to_excel(salvar, index=False)

        wb = load_workbook(salvar)
        ws = wb.active
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for i, encontrado in enumerate(df['Encontrado'], start=2):
            if encontrado:
                for cell in ws[i]:
                    cell.fill = fill_red
            if i % 5 == 0:
                progresso_janela.after(0, lambda v=90 + (i / total) * 10: atualizar_progresso(v))

        wb.save(salvar)

        progresso_janela.after(0, lambda: [
            progresso_janela.destroy(),
            messagebox.showinfo(
                "Concluído",
                f"✅ Análise concluída!\n\n"
                f"Total de jogadas: {total}\n"
                f"Pares encontrados: {encontrados_total} ({percentual_encontrado:.2f}%)\n"
                f"Acertos: {acertos} | Erros: {erros}\n"
                f"Taxa de acerto total: {percentual_acerto:.2f}%\n\n"
                f"Arquivo salvo em:\n{salvar}"
            )
        ])

    threading.Thread(target=executar, daemon=True).start()
    progresso_janela.mainloop()

root = tk.Tk()
root.title("Analisador de Probabilidades Blaze")
root.geometry("520x280")
root.configure(bg="#f8f8f8")

tk.Label(root, text="🎯 Analisador Blaze (Probabilidade 100/50 + Resultado Automático)",
         font=("Arial", 12, "bold"), bg="#f8f8f8", wraplength=400, justify="center").pack(pady=20)

tk.Button(root, text="📂 Selecionar e Analisar Planilha", command=analisar_planilha,
          font=("Arial", 12, "bold"), bg="#4CAF50", fg="white", padx=20, pady=10).pack(pady=15)

tk.Label(root, text="Linhas encontradas na lista serão destacadas em vermelho.\n"
                    "Resultado: Acerto G0, Acerto G1 ou Erro.",
         bg="#f8f8f8", font=("Arial", 10)).pack(pady=(10, 0))

root.mainloop()
